import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PDF_PATH = r"C:\Users\SamuelJoshuaRaj\Downloads\2021 llgudi poll data.pdf"   # ← change to your PDF path
OUTPUT_PATH = r"C:\Users\SamuelJoshuaRaj\OneDrive - CYGNUSA Technologies\Desktop\lalgudi pdf\lalgudi_poll_2021.xlsx"

candidates = [
    "A.SOUNDARAPANDIAN", "D.R.DHARMARAJ", "K.KAMARAJ", "R.SILAMBARASAN",
    "P.NAMBIRAJAN", "I.MALAR TAMIL PRABHA", "VEERAN.MUTHUKUMAR",
    "M.VIJAYAMURTHY", "K.ANANTHABABU", "A.ANANDHKUMAR", "P.M.SAHADEVAN",
    "ANBIL K.THANGAMANI", "K.DHARMARAJ", "U.JOHNSON"
]

parties = [
    "Dravida Munnetra Kazhagam",
    "All India Anna Dravida Munnetra Kazhagam",
    "Samaniya Makkal Nala Katchi",
    "Anna MGR Dravida Makkal Kalgam",
    "Puthiya Tamilagam",
    "Naam Tamilar Katchi",
    "Shiva Sena",
    "Amma Makkal Munnettra Kazagam",
    "Independent", "Independent", "Independent",
    "Independent", "Independent", "Independent"
]

def style_cell(cell, bold=False, bg=None, font_color="000000", size=9, wrap=True, align="center"):
    cell.font = Font(name="Arial", bold=bold, color=font_color, size=size)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    thin = Side(style="thin", color="AAAAAA")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def extract_rows_from_pdf(pdf_path):
    """Extract numeric data rows from the PDF using pdfplumber."""
    all_rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            for line in text.split("\n"):
                tokens = line.strip().split()
                # A valid data row starts with a serial number (integer)
                # followed by a polling station id, then 14 numeric votes + total + rejected
                if not tokens:
                    continue
                try:
                    int(tokens[0])          # serial no must be integer
                except ValueError:
                    continue
                # Collect all numeric tokens after the serial number
                nums = []
                ps_id = None
                for t in tokens[1:]:
                    # polling station id may be like "1", "1(A)", "23(A)" etc.
                    if ps_id is None:
                        ps_id = t
                        continue
                    try:
                        nums.append(int(t))
                    except ValueError:
                        pass  # skip non-numeric tokens (header text leaking in)
                # We expect 16 numbers: 14 candidate votes + total valid + rejected
                if len(nums) == 16:
                    all_rows.append((int(tokens[0]), ps_id, *nums))
    return all_rows

def build_excel(rows, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Poll Data"

    # ── Row 1: Title ──────────────────────────────────────────────────
    ws.merge_cells("A1:R1")
    ws["A1"] = "143 - Lalgudi Assembly Election, TNLA-2021  |  Total Electors: 218131"
    style_cell(ws["A1"], bold=True, bg="1F4E79", font_color="FFFFFF", size=12)
    ws.row_dimensions[1].height = 24

    # ── Row 2: Column headers ─────────────────────────────────────────
    headers = ["Serial No.", "Polling Station No."] + candidates + ["Total Valid Votes", "No. of Rejected Votes"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        style_cell(c, bold=True, bg="2E75B6", font_color="FFFFFF")
    ws.row_dimensions[2].height = 60

    # ── Row 3: Party names ────────────────────────────────────────────
    party_row = ["", ""] + parties + ["", ""]
    for col, p in enumerate(party_row, 1):
        c = ws.cell(row=3, column=col, value=p)
        style_cell(c, bold=False, bg="375F8A", font_color="FFFFFF", size=8)
    ws.row_dimensions[3].height = 52

    # ── Data rows ─────────────────────────────────────────────────────
    for i, row in enumerate(rows):
        r = 4 + i
        bg = "EBF3FB" if i % 2 == 0 else "FFFFFF"
        for col, val in enumerate(row, 1):
            c = ws.cell(row=r, column=col, value=val)
            style_cell(c, bg=bg, align="left" if col == 2 else "center")

    # ── Summary rows (last 3 rows of PDF) ────────────────────────────
    last = 4 + len(rows)

    # Total Votes (booth)
    booth   = [83264,67612,268,138,123,16134,64,2913,56,92,65,246,120,303,171398]
    ws.merge_cells(f"A{last}:B{last}")
    ws.cell(row=last, column=1, value="Total Votes")
    style_cell(ws.cell(row=last, column=1), bold=True, bg="FFF2CC")
    for col, v in enumerate(booth, 3):
        style_cell(ws.cell(row=last, column=col, value=v), bold=True, bg="FFF2CC")

    # Postal Ballot
    postal  = [1650,353,2,0,1,114,1,28,14,0,1,0,2,1,2156]
    ws.merge_cells(f"A{last+1}:B{last+1}")
    ws.cell(row=last+1, column=1, value="Total Postal Ballot Votes")
    style_cell(ws.cell(row=last+1, column=1), bold=True, bg="E2EFDA")
    for col, v in enumerate(postal, 3):
        style_cell(ws.cell(row=last+1, column=col, value=v), bold=True, bg="E2EFDA")

    # Grand Total
    grand   = [84914,67965,270,138,124,16248,65,2941,57,96,65,247,120,304,173554]
    ws.merge_cells(f"A{last+2}:B{last+2}")
    ws.cell(row=last+2, column=1, value="Total Votes Polled")
    style_cell(ws.cell(row=last+2, column=1), bold=True, bg="1F4E79", font_color="FFFFFF")
    for col, v in enumerate(grand, 3):
        style_cell(ws.cell(row=last+2, column=col, value=v), bold=True, bg="1F4E79", font_color="FFFFFF")

    # ── Column widths & freeze ────────────────────────────────────────
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    for col in range(3, 17):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    ws.column_dimensions["Q"].width = 16
    ws.column_dimensions["R"].width = 14
    ws.freeze_panes = "C4"

    wb.save(output_path)
    print(f"✅ Saved → {output_path}  ({len(rows)} data rows extracted)")

if __name__ == "__main__":
    print("📄 Reading PDF...")
    rows = extract_rows_from_pdf(PDF_PATH)
    print(f"   Found {len(rows)} polling station rows")
    print("📊 Building Excel...")
    build_excel(rows, OUTPUT_PATH)